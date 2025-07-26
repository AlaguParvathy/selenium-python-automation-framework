import softest
import logging
import inspect
from openpyxl import Workbook, load_workbook
import csv


class Utils(softest.TestCase):
    def assert_list_item_text(self,element_list,desired_value):
        for stop in element_list:
            print("The text is: " + stop.text)
            #here 'self.assertEqual' is an argument that specifies what were willing to do
            #since we are checking if the values are equal, we use the same
            self.soft_assert(self.assertEqual,stop.text,desired_value)
            #the code line before optimization
            #assert stop.text == desired_value
            #to print something on the console
            if stop.text == desired_value:
                print("assert pass")
            else:
                print("assert fail")
            self.assert_all()

    def custom_logger(logLevel = logging.DEBUG):
        #this logger_name will help use identify which script the log is coming from
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        #the log file will be created in Utilities package
        #by default the file mode is append
        fh = logging.FileHandler('automation.log', mode='a')
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', datefmt= '%m/%d/%Y %I:%M:%S %p')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet_name):
        datalist = []
        wb = load_workbook(filename = file_name)
        sh = wb[sheet_name]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for row in range(2,row_ct+1):
            row_data = []
            for col in range(1,col_ct+1):
                row_data.append(sh.cell(row = row, column = col).value)
            datalist.append(row_data)

        return datalist

    def read_data_from_csv(file_name):
        datalist = []
        #open the csv file
        csvdata = open(file_name, 'r')
        #create a csv reader
        reader = csv.reader(csvdata)
        #skip the header row
        next(reader)
        #add csv rows to list
        for row in reader:
            datalist.append(row)
        return datalist




